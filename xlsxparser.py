import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet import worksheet
from loguru import logger
from math import floor
from .utils import int_try_parse

class Parser():

    wb: Workbook
    sheet: worksheet

    def __init__(self, excel_file: str, start_sheet: str) -> None:
        self.wb = openpyxl.load_workbook(excel_file, data_only=True, keep_links=True)
        self.sheet = self.wb[start_sheet]

    def get_cell_value_from_cell(self, cell: str) -> str:
        return self.sheet[cell].value 
    
    def get_cell_value_from_rc(self, row: int, col: int) -> str:
        return self.sheet.cell(row=row, column=col).value
    
    def set_sheet(self, sheet_name: str):
        self.sheet = self.wb[sheet_name]
        
    def get_range_from_cells(self, range_start: str, range_end: str): #range_end указывается не строго [range_start, range_end)
        range = self.sheet[f'{range_start}:{range_end}']
        res = []
        for cells in range:
            for cell in cells:
                value = cell.value
                res.append(str(value))
        return res 
    
    def get_range_from_rc(self, start_row: int, start_col: int, end_row: int, end_col: int): #range_end указывается не строго [range_start, range_end)
        res = []
        for cells in self.sheet.iter_rows(min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col):
            for cell in cells:
                res.append(str(cell.value))
        return res 

    def get_range_top_untill(self, start_cell: str, end_cell: str, separator):
        res = []
        start_letter = str(self.split_cell(start_cell)[0])
        start_num = int(self.split_cell(start_cell)[1])
        end_letter = str(self.split_cell(end_cell)[0])
        end_num = int(self.split_cell(end_cell)[1])

        if type(separator) == int:
            step_num = end_num - start_num + 1
            res.append(self.get_range_from_cells(start_cell, end_cell))

            for _ in range(separator - 1):
                start_num = start_num - step_num
                end_num = end_num - step_num
                if start_num <= 0: return res
                next_start_cell = f'{start_letter}{start_num}'
                next_end_cell = f'{end_letter}{end_num}'
                res.append(self.get_range_from_cells(next_start_cell, next_end_cell))

            return res
                
        elif type(separator) == str:
            step_num = end_num - start_num + 1
            range_value = self.get_range_from_cells(start_cell, end_cell)
            if separator in(range_value): return res
            res.append(range_value)
            while True:
                start_num = start_num - step_num
                end_num = end_num - step_num
                next_start_cell = f'{start_letter}{start_num}'
                next_end_cell = f'{end_letter}{end_num}'
                if start_num <= 0: return res
                range_value = self.get_range_from_cells(next_start_cell, next_end_cell)
                if separator in(range_value): return res
                res.append(range_value)

            return res
        
        elif separator == None:
            step_num = end_num - start_num + 1

            data_cell = start_cell
            data_cell_value = self.get_cell_value_from_cell(data_cell)
            if data_cell_value == separator: return res

            range_value = self.get_range_from_cells(start_cell, end_cell)
            res.append(range_value)
            while True:
                start_num = start_num - step_num
                end_num = end_num - step_num
                next_start_cell = f'{start_letter}{start_num}'
                next_end_cell = f'{end_letter}{end_num}'
                if start_num <= 0: return res

                data_cell = next_start_cell
                data_cell_value = self.get_cell_value_from_cell(data_cell)
                if data_cell_value == separator: return res

                range_value = self.get_range_from_cells(next_start_cell, next_end_cell)
                res.append(range_value)
            return res
        else:
            raise Exception(f"Необрабатываемый тип {type(separator)}")

    def get_range_bottom_untill(self, start_cell: str, end_cell: str, separator):
        res = []
        start_letter = str(self.split_cell(start_cell)[0])
        start_num = int(self.split_cell(start_cell)[1])
        end_letter = str(self.split_cell(end_cell)[0])
        end_num = int(self.split_cell(end_cell)[1])

        if type(separator) == int:
            step_num = end_num - start_num + 1
            res.append(self.get_range_from_cells(start_cell, end_cell))

            for _ in range(separator - 1):
                start_num = start_num + step_num
                end_num = end_num + step_num
                if start_num <= 0: return res
                next_start_cell = f'{start_letter}{start_num}'
                next_end_cell = f'{end_letter}{end_num}'
                res.append(self.get_range_from_cells(next_start_cell, next_end_cell))

            return res
                
        elif type(separator) == str:
            step_num = end_num - start_num + 1
            range_value = self.get_range_from_cells(start_cell, end_cell)
            if separator in(range_value): return res
            res.append(range_value)
            while True:
                start_num = start_num + step_num
                end_num = end_num + step_num
                next_start_cell = f'{start_letter}{start_num}'
                next_end_cell = f'{end_letter}{end_num}'
                if start_num <= 0: return res
                range_value = self.get_range_from_cells(next_start_cell, next_end_cell)
                if separator in(range_value): return res
                res.append(range_value)

            return res
        
        elif separator == None:
            step_num = end_num - start_num + 1

            data_cell = start_cell
            data_cell_value = self.get_cell_value_from_cell(data_cell)
            if data_cell_value == separator: return res

            range_value = self.get_range_from_cells(start_cell, end_cell)
            res.append(range_value)
            while True:
                start_num = start_num + step_num
                end_num = end_num + step_num
                next_start_cell = f'{start_letter}{start_num}'
                next_end_cell = f'{end_letter}{end_num}'
                if start_num <= 0: return res

                data_cell = next_start_cell
                data_cell_value = self.get_cell_value_from_cell(data_cell)
                if data_cell_value == separator: return res

                range_value = self.get_range_from_cells(next_start_cell, next_end_cell)
                res.append(range_value)
            return res
        else:
            raise Exception(f"Необрабатываемый тип {type(separator)}")
        
    def get_range_left_untill(self, start_cell: str, end_cell: str, separator):
        res = []

        start_cell_row = self.sheet[start_cell].row
        start_cell_col = self.sheet[start_cell].column
        end_cell_row = self.sheet[end_cell].row
        end_cell_col = self.sheet[end_cell].column

        if type(separator) == int:
            res.append(self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col))
            for _ in range(separator - 1):
                start_cell_col = start_cell_col - 1
                end_cell_col = end_cell_col - 1
                if start_cell_col <= 0: return res
                res.append(self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col))

            return res
                
        elif type(separator) == str:
            range_value = self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col)
            if separator in(range_value): return res
            res.append(range_value)
            while True:
                start_cell_col = start_cell_col - 1
                end_cell_col = end_cell_col - 1
                if start_cell_col <= 0: return res
                range_value = self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col)
                if separator in(range_value): return res
                res.append(range_value)

            return res
        
        elif separator == None:
            
            data_cell_value = self.get_cell_value_from_rc(start_cell_row, start_cell_col)
            if data_cell_value == separator: return res
            range_value = self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col)
            res.append(range_value)
            while True:
                start_cell_col = start_cell_col - 1
                end_cell_col = end_cell_col - 1

                data_cell_value = self.get_cell_value_from_rc(start_cell_row, start_cell_col)
                if data_cell_value == separator: return res
                if start_cell_col <= 0: return res

                range_value = self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col)
                res.append(range_value)
        else:
            raise Exception(f"Необрабатываемый тип {type(separator)}")
        
    def get_range_right_untill(self, start_cell: str, end_cell: str, separator):
        res = []

        start_cell_row = self.sheet[start_cell].row
        start_cell_col = self.sheet[start_cell].column
        end_cell_row = self.sheet[end_cell].row
        end_cell_col = self.sheet[end_cell].column

        if type(separator) == int:
            res.append(self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col))
            for _ in range(separator - 1):
                start_cell_col = start_cell_col + 1
                end_cell_col = end_cell_col + 1
                if start_cell_col <= 0: return res
                res.append(self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col))

            return res
                
        elif type(separator) == str:
            range_value = self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col)
            if separator in(range_value): return res
            res.append(range_value)
            while True:
                start_cell_col = start_cell_col + 1
                end_cell_col = end_cell_col + 1
                if start_cell_col <= 0: return res
                range_value = self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col)
                if separator in(range_value): return res
                res.append(range_value)

            return res
        
        elif separator == None:
            
            data_cell_value = self.get_cell_value_from_rc(start_cell_row, start_cell_col)
            if data_cell_value == separator: return res
            range_value = self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col)
            res.append(range_value)
            while True:
                start_cell_col = start_cell_col + 1
                end_cell_col = end_cell_col + 1

                data_cell_value = self.get_cell_value_from_rc(start_cell_row, start_cell_col)
                if data_cell_value == separator: return res
                if start_cell_col <= 0: return res

                range_value = self.get_range_from_rc(start_cell_row, start_cell_col, end_cell_row, end_cell_col)
                res.append(range_value)
        else:
            raise Exception(f"Необрабатываемый тип {type(separator)}")
    

    #TODO Заменить разбор ячейки со срезов на этот метод
    def split_cell(self, cell: str):
        leng = len(cell)
        middle_index = floor(leng / 2) - 1
        num_index = 0
        value = int_try_parse(cell[middle_index])
        while type(value) == int:
            middle_index -= 1
            value = int_try_parse(cell[middle_index])
        num_index = middle_index + 1
        value = int_try_parse(cell[num_index])
        while type(value) != int:
            num_index += 1
            value = int_try_parse(cell[num_index])

        num = cell[num_index:]
        letter = cell[:num_index]

        return (letter, num)